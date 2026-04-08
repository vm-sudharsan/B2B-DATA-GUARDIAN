"""
Professional Excel Report Generator with Multiple Sheets and Formatting
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
import pandas as pd
from datetime import datetime
from io import BytesIO


class ProfessionalExcelReport:
    """Generate professional Excel reports with multiple sheets and formatting"""
    
    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet
        
        # Define color scheme
        self.colors = {
            'header': 'FF4F46E5',  # Indigo
            'subheader': 'FF818CF8',  # Light Indigo
            'success': 'FF10B981',  # Green
            'warning': 'FFF59E0B',  # Amber
            'error': 'FFEF4444',  # Red
            'info': 'FF3B82F6',  # Blue
            'light_gray': 'FFF3F4F6',
            'white': 'FFFFFFFF'
        }
        
        # Define fonts
        self.fonts = {
            'title': Font(name='Calibri', size=16, bold=True, color='FFFFFFFF'),
            'header': Font(name='Calibri', size=12, bold=True, color='FFFFFFFF'),
            'subheader': Font(name='Calibri', size=11, bold=True),
            'normal': Font(name='Calibri', size=10),
            'small': Font(name='Calibri', size=9)
        }
        
        # Define borders
        thin_border = Side(style='thin', color='FFD1D5DB')
        self.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    
    def generate_report(self, original_data, cleaned_data, report_metrics, fixes, 
                       missing_records, invalid_records, duplicate_records, output_path=None):
        """Generate comprehensive Excel report"""
        
        # Sheet 1: Summary Dashboard
        self._create_summary_sheet(report_metrics, len(cleaned_data))
        
        # Sheet 2: Cleaned Data
        self._create_cleaned_data_sheet(cleaned_data)
        
        # Sheet 3: Data Quality Issues
        self._create_issues_sheet(missing_records, invalid_records)
        
        # Sheet 4: Duplicates Analysis
        self._create_duplicates_sheet(duplicate_records, cleaned_data)
        
        # Sheet 5: Changes Log
        self._create_changes_sheet(fixes)
        
        # Sheet 6: Field Analysis
        self._create_field_analysis_sheet(cleaned_data, report_metrics)
        
        # Save or return
        if output_path:
            self.wb.save(output_path)
            return output_path
        else:
            output = BytesIO()
            self.wb.save(output)
            output.seek(0)
            return output
    
    def _create_summary_sheet(self, metrics, total_records):
        """Create executive summary dashboard"""
        ws = self.wb.create_sheet("📊 Summary Dashboard", 0)
        
        # Title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = "Data Quality Analysis Report"
        title_cell.font = self.fonts['title']
        title_cell.fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Subtitle
        ws.merge_cells('A2:F2')
        subtitle_cell = ws['A2']
        subtitle_cell.value = f"Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        subtitle_cell.font = self.fonts['small']
        subtitle_cell.alignment = Alignment(horizontal='center')
        
        # Quality Score Section
        ws.merge_cells('A4:C4')
        ws['A4'] = "Overall Quality Score"
        ws['A4'].font = self.fonts['subheader']
        ws['A4'].fill = PatternFill(start_color=self.colors['light_gray'], end_color=self.colors['light_gray'], fill_type='solid')
        
        quality_score = metrics.get('overall_quality_score', 0)
        ws.merge_cells('A5:C5')
        score_cell = ws['A5']
        score_cell.value = f"{quality_score:.1f}%"
        score_cell.font = Font(name='Calibri', size=36, bold=True, color=self._get_score_color(quality_score))
        score_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[5].height = 50
        
        ws.merge_cells('A6:C6')
        ws['A6'] = self._get_quality_label(quality_score)
        ws['A6'].font = Font(name='Calibri', size=14, bold=True)
        ws['A6'].alignment = Alignment(horizontal='center')
        
        # Key Metrics
        row = 8
        metrics_data = [
            ("Total Records", total_records, self.colors['info']),
            ("Duplicates Found", metrics.get('duplicate_rows', 0), self.colors['warning']),
            ("Issues Fixed", metrics.get('offline_fixes', 0) + metrics.get('online_fixes', 0), self.colors['success']),
            ("Manual Review", metrics.get('manual_review', 0), self.colors['error']),
            ("Missing Fields", metrics.get('missing_fields', 0), self.colors['warning']),
            ("Invalid Formats", metrics.get('invalid_fields', 0), self.colors['error']),
        ]
        
        for label, value, color in metrics_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = self.fonts['normal']
            ws[f'A{row}'].border = self.border
            
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = Font(name='Calibri', size=12, bold=True)
            ws[f'B{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            ws[f'B{row}'].font = Font(name='Calibri', size=12, bold=True, color='FFFFFFFF')
            ws[f'B{row}'].alignment = Alignment(horizontal='center')
            ws[f'B{row}'].border = self.border
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
    
    def _create_cleaned_data_sheet(self, cleaned_data):
        """Create cleaned data sheet with professional formatting"""
        ws = self.wb.create_sheet("✅ Cleaned Data")
        
        # Add data
        for r_idx, row in enumerate(dataframe_to_rows(cleaned_data, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = self.border
                
                if r_idx == 1:  # Header row
                    cell.font = self.fonts['header']
                    cell.fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.font = self.fonts['normal']
                    if r_idx % 2 == 0:
                        cell.fill = PatternFill(start_color=self.colors['light_gray'], end_color=self.colors['light_gray'], fill_type='solid')
        
        # Auto-adjust column widths
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    if cell.value is not None:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    def _create_issues_sheet(self, missing_records, invalid_records):
        """Create data quality issues sheet"""
        ws = self.wb.create_sheet("⚠️ Quality Issues")
        
        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = "Data Quality Issues"
        ws['A1'].font = self.fonts['title']
        ws['A1'].fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        # Missing Fields Section
        row = 3
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = f"Missing Fields ({len(missing_records)} issues)"
        ws[f'A{row}'].font = self.fonts['subheader']
        ws[f'A{row}'].fill = PatternFill(start_color=self.colors['warning'], end_color=self.colors['warning'], fill_type='solid')
        ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color='FFFFFFFF')
        
        row += 1
        headers = ['Row #', 'Field', 'Issue', 'Severity', 'Recommendation']
        for c_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c_idx, value=header)
            cell.font = self.fonts['header']
            cell.fill = PatternFill(start_color=self.colors['subheader'], end_color=self.colors['subheader'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        row += 1
        for record in missing_records[:100]:  # Limit to 100
            ws.cell(row=row, column=1, value=record.get('row_index', '') + 1)
            ws.cell(row=row, column=2, value=record.get('field', ''))
            ws.cell(row=row, column=3, value=record.get('issue', ''))
            ws.cell(row=row, column=4, value='High')
            ws.cell(row=row, column=5, value='Fill missing data')
            
            for c_idx in range(1, 6):
                ws.cell(row=row, column=c_idx).border = self.border
                ws.cell(row=row, column=c_idx).font = self.fonts['normal']
            row += 1
        
        # Invalid Formats Section
        row += 2
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = f"Invalid Formats ({len(invalid_records)} issues)"
        ws[f'A{row}'].font = self.fonts['subheader']
        ws[f'A{row}'].fill = PatternFill(start_color=self.colors['error'], end_color=self.colors['error'], fill_type='solid')
        ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color='FFFFFFFF')
        
        row += 1
        headers = ['Row #', 'Field', 'Invalid Value', 'Issue', 'Recommendation']
        for c_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c_idx, value=header)
            cell.font = self.fonts['header']
            cell.fill = PatternFill(start_color=self.colors['subheader'], end_color=self.colors['subheader'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        row += 1
        for record in invalid_records[:100]:  # Limit to 100
            ws.cell(row=row, column=1, value=record.get('row_index', '') + 1)
            ws.cell(row=row, column=2, value=record.get('field', ''))
            ws.cell(row=row, column=3, value=str(record.get('value', ''))[:50])
            ws.cell(row=row, column=4, value=record.get('issue', ''))
            ws.cell(row=row, column=5, value='Correct format')
            
            for c_idx in range(1, 6):
                ws.cell(row=row, column=c_idx).border = self.border
                ws.cell(row=row, column=c_idx).font = self.fonts['normal']
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 25
    
    def _create_duplicates_sheet(self, duplicate_records, cleaned_data):
        """Create duplicates analysis sheet"""
        ws = self.wb.create_sheet("🔄 Duplicates")
        
        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = f"Duplicate Records Analysis ({len(duplicate_records)} duplicates found)"
        ws['A1'].font = self.fonts['title']
        ws['A1'].fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        if duplicate_records:
            # Get columns from cleaned data
            columns = list(cleaned_data.columns)
            
            # Headers
            row = 3
            ws.cell(row=row, column=1, value='Duplicate Group')
            ws.cell(row=row, column=1).font = self.fonts['header']
            ws.cell(row=row, column=1).fill = PatternFill(start_color=self.colors['subheader'], end_color=self.colors['subheader'], fill_type='solid')
            
            for c_idx, col in enumerate(columns, 2):
                cell = ws.cell(row=row, column=c_idx, value=col)
                cell.font = self.fonts['header']
                cell.fill = PatternFill(start_color=self.colors['subheader'], end_color=self.colors['subheader'], fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
                cell.border = self.border
            
            # Data
            row += 1
            group_num = 1
            for dup in duplicate_records[:100]:  # Limit to 100
                row_idx = dup.get('row_index', 0)
                if row_idx < len(cleaned_data):
                    ws.cell(row=row, column=1, value=f"Group {group_num}")
                    ws.cell(row=row, column=1).fill = PatternFill(start_color=self.colors['warning'], end_color=self.colors['warning'], fill_type='solid')
                    ws.cell(row=row, column=1).font = Font(name='Calibri', size=10, bold=True)
                    
                    record = cleaned_data.iloc[row_idx]
                    for c_idx, col in enumerate(columns, 2):
                        value = record[col] if col in record else ''
                        ws.cell(row=row, column=c_idx, value=str(value)[:50])
                        ws.cell(row=row, column=c_idx).border = self.border
                        ws.cell(row=row, column=c_idx).font = self.fonts['normal']
                    
                    row += 1
                    group_num += 1
        else:
            ws['A3'] = "✓ No duplicate records found. Your data is clean!"
            ws['A3'].font = Font(name='Calibri', size=14, bold=True, color='FF10B981')
        
        # Auto-adjust column widths
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    if cell.value is not None:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_changes_sheet(self, fixes):
        """Create changes log sheet"""
        ws = self.wb.create_sheet("📝 Changes Log")
        
        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = f"Data Changes Log ({len(fixes)} changes applied)"
        ws['A1'].font = self.fonts['title']
        ws['A1'].fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        # Headers
        row = 3
        headers = ['Row #', 'Field', 'Original Value', 'Corrected Value', 'Confidence', 'Mode']
        for c_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c_idx, value=header)
            cell.font = self.fonts['header']
            cell.fill = PatternFill(start_color=self.colors['subheader'], end_color=self.colors['subheader'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Data
        row += 1
        for fix in fixes[:500]:  # Limit to 500
            ws.cell(row=row, column=1, value=fix.get('row_index', '') + 1)
            ws.cell(row=row, column=2, value=fix.get('field', ''))
            ws.cell(row=row, column=3, value=str(fix.get('original', ''))[:50])
            ws.cell(row=row, column=4, value=str(fix.get('suggested', ''))[:50])
            ws.cell(row=row, column=5, value=fix.get('confidence', 0))
            ws.cell(row=row, column=6, value=fix.get('processing_mode', ''))
            
            # Color code by mode
            mode = str(fix.get('processing_mode', '')).lower()
            if mode == 'accept':
                fill_color = self.colors['success']
            elif mode == 'suggest':
                fill_color = self.colors['warning']
            elif mode == 'manual':
                fill_color = self.colors['error']
            else:
                fill_color = self.colors['info']
            
            ws.cell(row=row, column=6).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            ws.cell(row=row, column=6).font = Font(name='Calibri', size=10, bold=True, color='FFFFFFFF')
            ws.cell(row=row, column=6).alignment = Alignment(horizontal='center')
            
            for c_idx in range(1, 7):
                ws.cell(row=row, column=c_idx).border = self.border
                if c_idx != 6:
                    ws.cell(row=row, column=c_idx).font = self.fonts['normal']
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
    
    def _create_field_analysis_sheet(self, cleaned_data, metrics):
        """Create field-by-field analysis sheet"""
        ws = self.wb.create_sheet("📈 Field Analysis")
        
        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = "Field-by-Field Data Quality Analysis"
        ws['A1'].font = self.fonts['title']
        ws['A1'].fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        # Headers
        row = 3
        headers = ['Field Name', 'Total Values', 'Missing', 'Unique Values', 'Completeness %']
        for c_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c_idx, value=header)
            cell.font = self.fonts['header']
            cell.fill = PatternFill(start_color=self.colors['subheader'], end_color=self.colors['subheader'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Analyze each field
        row += 1
        for column in cleaned_data.columns:
            total = len(cleaned_data)
            missing = cleaned_data[column].isna().sum()
            unique = cleaned_data[column].nunique()
            completeness = ((total - missing) / total * 100) if total > 0 else 0
            
            ws.cell(row=row, column=1, value=column)
            ws.cell(row=row, column=2, value=total)
            ws.cell(row=row, column=3, value=missing)
            ws.cell(row=row, column=4, value=unique)
            ws.cell(row=row, column=5, value=f"{completeness:.1f}%")
            
            # Color code completeness
            if completeness >= 90:
                fill_color = self.colors['success']
            elif completeness >= 70:
                fill_color = self.colors['warning']
            else:
                fill_color = self.colors['error']
            
            ws.cell(row=row, column=5).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            ws.cell(row=row, column=5).font = Font(name='Calibri', size=10, bold=True, color='FFFFFFFF')
            ws.cell(row=row, column=5).alignment = Alignment(horizontal='center')
            
            for c_idx in range(1, 6):
                ws.cell(row=row, column=c_idx).border = self.border
                if c_idx != 5:
                    ws.cell(row=row, column=c_idx).font = self.fonts['normal']
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 18
    
    def _get_score_color(self, score):
        """Get color based on quality score"""
        if score >= 85:
            return 'FF10B981'  # Green
        elif score >= 70:
            return 'FF3B82F6'  # Blue
        elif score >= 50:
            return 'FFF59E0B'  # Orange
        else:
            return 'FFEF4444'  # Red
    
    def _get_quality_label(self, score):
        """Get quality label based on score"""
        if score >= 85:
            return "Excellent Quality"
        elif score >= 70:
            return "Good Quality"
        elif score >= 50:
            return "Fair Quality"
        else:
            return "Needs Improvement"
